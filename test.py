import openpyxl
import os

# 1️⃣ Excel-Dateien laden
quelle = openpyxl.load_workbook('Downloads/Pluxee Excel.xlsx')  # Datei, von der kopiert wird
ziel = openpyxl.load_workbook('Pluxee_Essenszuschuss_final.xlsx')      # Datei, in die eingefügt wird

# 2️⃣ Arbeitsblätter auswählen
ws_quelle = quelle.active
ws_ziel = ziel.active

start_zeile_quelle = 4
start_zeile_ziel = 9
for x in range(25):
    start_spalte_quelle = 1  # Spalte A
    end_spalte_quelle = 10   # Spalte J

    # Zielbereich
    start_spalte_ziel = 2    # Spalte B

    # 4️⃣ Werte kopieren
    for i, spalte in enumerate(range(start_spalte_quelle, end_spalte_quelle + 1)):
        wert = ws_quelle.cell(row=start_zeile_quelle, column=spalte).value
        ws_ziel.cell(row=start_zeile_ziel, column=start_spalte_ziel + i).value = wert

    # 5️⃣ Ziel-Excel speichern
    ziel.save('ziel.xlsx')
    start_zeile_quelle += 1
    start_zeile_ziel += 1

os.system("start ziel.xlsx")


